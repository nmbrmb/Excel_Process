import pandas as pd
import gradio as gr
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import time

def extract_birthdate_from_id(id_number):
    if isinstance(id_number, str) and len(id_number) == 18:
        year = int(id_number[6:10])
        month = int(id_number[10:12])
        day = int(id_number[12:14])
        return datetime(year, month, day)
    return None

def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

def process_files(original_file, new_files, company_order_file):
    # 读取公司名称顺序
    company_order_df = pd.read_excel(company_order_file)
    company_order = company_order_df['单位'].tolist()  # 假设公司名称列名为“单位”

    # 读取特定sheet名称的Excel表格，确保身份证号列作为字符串读取
    original_df = pd.read_excel(original_file.name, sheet_name='主动退出', dtype={'身份证号': str})
    original_df['出生年月（xx年xx月）'] = original_df['身份证号'].apply(
        lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
    original_df['年龄（有公式）'] = original_df['身份证号'].apply(
        lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")

    # 初始化一个 DataFrame 来存储所有新文件的数据
    new_df_list = []

    for file in new_files:
        temp_df = pd.read_excel(file.name, sheet_name='主动退出', dtype={'身份证号': str})
        temp_df['出生年月（xx年xx月）'] = temp_df['身份证号'].apply(
            lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
        temp_df['年龄（有公式）'] = temp_df['身份证号'].apply(
            lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")
        new_df_list.append(temp_df)

    # 合并所有新文件的数据
    combined_new_df = pd.concat(new_df_list, ignore_index=True)

    # 根据身份证号去重，仅保留新文件中的新增内容
    new_unique_df = combined_new_df[~combined_new_df['身份证号'].isin(original_df['身份证号'])]

    # 初始化最终的 DataFrame
    final_df = pd.DataFrame(columns=original_df.columns)

    # 遍历每个公司
    for company in company_order:
        # 获取原文件中当前公司的所有行
        original_company_df = original_df[original_df['单位'] == company]

        # 获取新文件中当前公司的新增行
        new_company_df = new_unique_df[new_unique_df['单位'] == company]

        # 将原公司的行和新公司的新增行按顺序添加到最终 DataFrame 中
        final_df = pd.concat([final_df, original_company_df, new_company_df], ignore_index=True)

    # 重新对“序号”列进行编号
    final_df['序号'] = range(1, len(final_df) + 1)

    # 保存到新的Excel文件，确保Sheet名称为“主动退出”
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_file = f'合并的排序表-生成时间({timestamp}).xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        final_df.to_excel(writer, index=False, sheet_name='主动退出')

    # 加载保存的Excel文件并标记新增行的所有内容为红色，缺失值的行标记为绿色
    workbook = load_workbook(output_file)
    sheet = workbook['主动退出']

    red_font = Font(color="FF0000")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # 找到新添加的行并标记
    for index, row in final_df.iterrows():
        is_new_row = row['身份证号'] in new_unique_df['身份证号'].values
        has_missing_value = pd.isnull(row['出生年月（xx年xx月）']) or row['出生年月（xx年xx月）'] == "" or pd.isnull(row['年龄（有公式）']) or row['年龄（有公式）'] == ""
        for col_num in range(1, len(row) + 1):
            cell = sheet.cell(row=index + 2, column=col_num)
            if is_new_row:
                cell.font = red_font
            if has_missing_value:
                cell.fill = green_fill

    # 保存修改后的Excel文件
    workbook.save(output_file)

    return output_file

def update_age_file(age_file):
    # 读取特定sheet名称的Excel表格，确保身份证号列作为字符串读取
    df = pd.read_excel(age_file.name, sheet_name='主动退出', dtype={'身份证号': str})

    # 提取出生年月和计算年龄
    df['出生年月（xx年xx月）'] = df['身份证号'].apply(
        lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
    df['年龄（有公式）'] = df['身份证号'].apply(
        lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")

    # 保存到新的Excel文件，确保Sheet名称为“主动退出”
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_file = f'年龄更新的排序表-生成时间({timestamp}).xlsx'
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='主动退出')

    return output_file

def get_current_time():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 创建Gradio界面
def interface():
    with gr.Blocks(title="Excel处理") as demo:
        with gr.Row():
            with gr.Column():
                original_file = gr.File(label="上传原文件")
                new_files = gr.Files(label="上传新文件（可多选）")
                company_order_file = gr.File(label="上传公司排序文件")
            with gr.Column():
                output = gr.File(label="下载合并排序后的文件")
        with gr.Row():
            instruction_text = gr.HTML("<b>新生成的文件中<span style='color: red;'>红色</span>代表新增的行，需要检查<b>")
        with gr.Row():
            instruction_text = gr.HTML(
                "<b>新生成的文件中<span style='color: green;'>绿色</span>代表“出生年月（xx年xx月）”或“年龄（有公式）”缺失值的行，需要检查身份证号，并填充缺失值</b>")
        btn = gr.Button("处理文件")

        with gr.Row():
            with gr.Column():
                age_file = gr.File(label="上传需要更新年龄的文件")
            with gr.Column():
                age_output = gr.File(label="下载更新后的文件")
        with gr.Row():
            current_time = gr.Textbox(label="当前时间", value=get_current_time(), interactive=False)
        age_btn = gr.Button("更新年龄")

        btn.click(fn=process_files, inputs=[original_file, new_files, company_order_file], outputs=output)
        age_btn.click(fn=update_age_file, inputs=[age_file], outputs=age_output)

        demo.load(get_current_time, inputs=None, outputs=current_time, every=1)

    demo.launch()

interface()