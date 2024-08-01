import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import time
import os

def extract_birthdate_from_id(id_number):
    if isinstance(id_number, str) and len(id_number) == 18:
        year = int(id_number[6:10])
        month = int(id_number[10:12])
        day = int(id_number[12:14])
        return datetime(year, month, day)
    return None

def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

def process_files(original_file, new_files, company_order_file):
    try:
        # 读取公司名称顺序
        company_order_df = pd.read_excel(company_order_file)
        company_order = company_order_df['单位'].tolist()  # 假设公司名称列名为“单位”

        # 读取特定sheet名称的Excel表格，确保身份证号列作为字符串读取
        original_df = pd.read_excel(original_file, sheet_name='主动退出', dtype={'身份证号': str})
        original_df['出生年月（xx年xx月）'] = original_df['身份证号'].apply(
            lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
        original_df['年龄（有公式）'] = original_df['身份证号'].apply(
            lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")

        # 初始化一个 DataFrame 来存储所有新文件的数据
        new_df_list = []

        for file in new_files:
            file = file.strip()  # 去除路径前后的空格
            if os.path.exists(file):  # 检查路径是否存在
                temp_df = pd.read_excel(file, sheet_name='主动退出', dtype={'身份证号': str})
                temp_df['出生年月（xx年xx月）'] = temp_df['身份证号'].apply(
                    lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
                temp_df['年龄（有公式）'] = temp_df['身份证号'].apply(
                    lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")
                new_df_list.append(temp_df)
            else:
                messagebox.showerror("错误", f"文件路径无效: {file}")
                return

        # 合并所有新文件的数据
        combined_new_df = pd.concat(new_df_list, ignore_index=True)

        # 根据身份证号去重，仅保留新文件中的新增内容
        new_unique_df = combined_new_df[~combined_new_df['身份证号'].isin(original_df['身份证号'])]

        # 初始化最终的 DataFrame
        final_df = pd.DataFrame(columns=original_df.columns)

        # 遍历每个公司
        for company in company_order:
            # 获取原文件中当前公司的所有行
            original_company_df = original_df[original_df['单位'] == company]

            # 获取新文件中当前公司的新增行
            new_company_df = new_unique_df[new_unique_df['单位'] == company]

            # 将原公司的行和新公司的新增行按顺序添加到最终 DataFrame 中
            final_df = pd.concat([final_df, original_company_df, new_company_df], ignore_index=True)

        # 重新对“序号”列进行编号
        final_df['序号'] = range(1, len(final_df) + 1)

        # 保存到新的Excel文件，确保Sheet名称为“主动退出”
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_file = f'合并的排序表-生成时间({timestamp}).xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            final_df.to_excel(writer, index=False, sheet_name='主动退出')

        # 加载保存的Excel文件并标记新增行的所有内容为红色，缺失值的行标记为绿色
        workbook = load_workbook(output_file)
        sheet = workbook['主动退出']

        red_font = Font(color="FF0000")
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # 找到新添加的行并标记
        for index, row in final_df.iterrows():
            is_new_row = row['身份证号'] in new_unique_df['身份证号'].values
            has_missing_value = pd.isnull(row['出生年月（xx年xx月）']) or row['出生年月（xx年xx月）'] == "" or pd.isnull(row['年龄（有公式）']) or row['年龄（有公式）'] == ""
            for col_num in range(1, len(row) + 1):
                cell = sheet.cell(row=index + 2, column=col_num)
                if is_new_row:
                    cell.font = red_font
                if has_missing_value:
                    cell.fill = green_fill

        # 保存修改后的Excel文件
        workbook.save(output_file)

        messagebox.showinfo("处理完成", f"文件已保存为 {output_file}")
    except Exception as e:
        messagebox.showerror("错误", f"处理文件时出错: {str(e)}")

def update_age_file(age_file):
    try:
        # 读取特定sheet名称的Excel表格，确保身份证号列作为字符串读取
        df = pd.read_excel(age_file, sheet_name='主动退出', dtype={'身份证号': str})

        # 提取出生年月和计算年龄
        df['出生年月（xx年xx月）'] = df['身份证号'].apply(
            lambda x: extract_birthdate_from_id(x).strftime('%Y年%m月') if extract_birthdate_from_id(x) else "")
        df['年龄（有公式）'] = df['身份证号'].apply(
            lambda x: calculate_age(extract_birthdate_from_id(x)) if extract_birthdate_from_id(x) else "")

        # 保存到新的Excel文件，确保Sheet名称为“主动退出”
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_file = f'年龄更新的排序表-生成时间({timestamp}).xlsx'
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='主动退出')

        messagebox.showinfo("处理完成", f"文件已保存为 {output_file}")
    except Exception as e:
        messagebox.showerror("错误", f"处理文件时出错: {str(e)}")

def get_current_time():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def select_original_file():
    original_file.set(filedialog.askopenfilename())

def select_new_files():
    new_files.set(";".join(filedialog.askopenfilenames()))

def select_company_order_file():
    company_order_file.set(filedialog.askopenfilename())

def select_age_file():
    age_file.set(filedialog.askopenfilename())

# 创建主窗口
root = tk.Tk()
root.title("Excel处理")

# 创建变量
original_file = tk.StringVar()
new_files = tk.StringVar()
company_order_file = tk.StringVar()
age_file = tk.StringVar()

# 创建并布局组件
tk.Label(root, text="上传原文件").grid(row=0, column=0)
tk.Entry(root, textvariable=original_file, width=50).grid(row=0, column=1)
tk.Button(root, text="选择文件", command=select_original_file).grid(row=0, column=2)

tk.Label(root, text="上传新文件（可多选）").grid(row=1, column=0)
tk.Entry(root, textvariable=new_files, width=50).grid(row=1, column=1)
tk.Button(root, text="选择文件", command=select_new_files).grid(row=1, column=2)

tk.Label(root, text="上传公司排序文件").grid(row=2, column=0)
tk.Entry(root, textvariable=company_order_file, width=50).grid(row=2, column=1)
tk.Button(root, text="选择文件", command=select_company_order_file).grid(row=2, column=2)

tk.Button(root, text="处理文件", command=lambda: process_files(original_file.get(), new_files.get().split(";"), company_order_file.get())).grid(row=3, column=1)

tk.Label(root, text="上传需要更新年龄的文件").grid(row=4, column=0)
tk.Entry(root, textvariable=age_file, width=50).grid(row=4, column=1)
tk.Button(root, text="选择文件", command=select_age_file).grid(row=4, column=2)

tk.Button(root, text="更新年龄", command=lambda: update_age_file(age_file.get())).grid(row=5, column=1)

tk.Label(root, text="当前时间").grid(row=6, column=0)
current_time_label = tk.Label(root, text=get_current_time())
current_time_label.grid(row=6, column=1)

def update_time():
    current_time_label.config(text=get_current_time())
    root.after(1000, update_time)

update_time()

# 启动主循环
root.mainloop()